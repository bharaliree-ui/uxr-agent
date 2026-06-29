"""
research_agent.py — A UX Research Synthesis Agent that outputs a linked Excel tracker.

WHAT IT DOES
    Point it at a folder of interview transcripts. It reads each one, clusters
    findings into themes, backs every theme with VERBATIM participant quotes,
    rates its confidence, and produces a 3-tab Excel workbook
    (Research_Report_Tracker.xlsx) where:
      - "Raw Data"  = every evidence quote, one per row
      - "Synthesis" = themes + quotes; each quote links back to its Raw Data row
      - "Report"    = stakeholder summary; each theme links to its evidence
    So every finding is one click away from the words that prove it.

WHY IT'S BUILT THIS WAY (the part you talk about in interviews)
    Naive "summarize these interviews" prompts hallucinate quotes, over-merge
    themes, and present guesses as findings. This agent is designed against
    those failures: verbatim attributed quotes, confidence ratings, a mandatory
    human-verification section, and traceability from every claim to its source.

    It's a true agent (a model in a loop with tools), not one big prompt: it
    decides which transcripts to read, reads them, then submits structured
    findings that get turned into the workbook.

RUN IT
    pip install -r requirements.txt        (anthropic + openpyxl)
    export ANTHROPIC_API_KEY="sk-ant-..."
    python research_agent.py
"""

import os
import sys
import glob
import json
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter


# Folder to read transcripts from. Defaults to "transcripts/", but you can point
# it elsewhere by passing a folder name:  python3 research_agent.py transcripts_experience
TRANSCRIPT_DIR = sys.argv[1] if len(sys.argv) > 1 else "transcripts"
OUTPUT_FILE = "Research_Report_Tracker.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# 1. TOOLS
# ─────────────────────────────────────────────────────────────────────────────

def list_transcripts() -> str:
    files = sorted(os.path.basename(p) for p in glob.glob(f"{TRANSCRIPT_DIR}/*.txt"))
    return "\n".join(files) if files else "No transcripts found."


def read_transcript(filename: str) -> str:
    path = os.path.join(TRANSCRIPT_DIR, os.path.basename(filename))
    if not os.path.exists(path):
        return f"Error: {filename} not found."
    with open(path) as f:
        return f.read()


def submit_findings(product: str, participants: list, themes: list,
                    verification_notes: list) -> str:
    """Receive the agent's structured synthesis and build the Excel tracker."""
    data = {
        "product": product,
        "participants": participants,
        "themes": themes,
        "verification_notes": verification_notes,
    }
    build_workbook(data, OUTPUT_FILE)
    n_quotes = sum(len(t.get("quotes", [])) for t in themes)
    return (f"Built {OUTPUT_FILE}: {len(themes)} themes, {n_quotes} linked quotes, "
            f"{len(participants)} participants.")


TOOL_FUNCTIONS = {
    "list_transcripts": list_transcripts,
    "read_transcript": read_transcript,
    "submit_findings": submit_findings,
}


# ─────────────────────────────────────────────────────────────────────────────
# 2. TOOL MENU
# ─────────────────────────────────────────────────────────────────────────────

TOOLS = [
    {
        "name": "list_transcripts",
        "description": "List every interview transcript filename available. Call this first.",
        "input_schema": {"type": "object", "properties": {}},
    },
    {
        "name": "read_transcript",
        "description": "Read the full text of ONE transcript. Read every transcript before synthesizing.",
        "input_schema": {
            "type": "object",
            "properties": {"filename": {"type": "string", "description": "e.g. 'P1.txt'"}},
            "required": ["filename"],
        },
    },
    {
        "name": "submit_findings",
        "description": ("Submit the final synthesis as structured data. Call ONCE at the end. "
                        "This builds the Excel report. Every quote must be verbatim and attributed."),
        "input_schema": {
            "type": "object",
            "properties": {
                "product": {"type": "string", "description": "Name of the product studied"},
                "participants": {
                    "type": "array", "items": {"type": "string"},
                    "description": "Participant IDs, e.g. ['P1','P2','P3','P4']",
                },
                "themes": {
                    "type": "array",
                    "description": "3-6 themes, each with at least two supporting quotes from different participants",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string"},
                            "confidence": {"type": "string", "enum": ["High", "Medium", "Low"]},
                            "summary": {"type": "string", "description": "2-3 sentence summary"},
                            "recommendation": {"type": "string", "description": "One concrete, testable suggestion"},
                            "quotes": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "participant": {"type": "string"},
                                        "topic": {"type": "string", "description": "Short tag, e.g. 'Onboarding'"},
                                        "quote": {"type": "string", "description": "VERBATIM quote, no paraphrasing"},
                                    },
                                    "required": ["participant", "quote"],
                                },
                            },
                        },
                        "required": ["name", "confidence", "summary", "recommendation", "quotes"],
                    },
                },
                "verification_notes": {
                    "type": "array", "items": {"type": "string"},
                    "description": "Things a human must verify: sample caveats, thin evidence, inferences, open questions",
                },
            },
            "required": ["product", "participants", "themes", "verification_notes"],
        },
    },
]


SYSTEM_PROMPT = """You are a UX research synthesis assistant. You turn interview \
transcripts into a structured, evidence-driven insights draft for a human researcher.

PROCESS
1. Call list_transcripts.
2. Read EVERY transcript, one at a time, before drawing conclusions.
3. Cluster findings into 3-6 themes.
4. Call submit_findings ONCE with the structured result.

RULES THAT KEEP YOU HONEST
- Quotes must be VERBATIM and attributed to the right participant. Never invent or \
paraphrase a quote. If you can't find a real supporting quote, you don't have the theme.
- Each theme needs at least TWO quotes from DIFFERENT participants.
- Rate each theme's confidence High / Medium / Low by how many participants raised it \
and how consistent they were.
- verification_notes MUST list sample limitations, inferences, thin evidence, and open \
questions. This is a draft for a researcher to check, never a final answer.

After submit_findings succeeds, give the user a short plain-language summary and tell \
them to open the Excel file."""


# ─────────────────────────────────────────────────────────────────────────────
# 3. WORKBOOK BUILDER — turns structured findings into the 3-tab linked tracker.
# ─────────────────────────────────────────────────────────────────────────────

ARIAL = "Arial"
NAVY, HEADER, GREEN, AMBER, LINK = "1F3864", "2E5496", "C6EFCE", "FFEB9C", "0563C1"
_thin = Side(style="thin", color="BFBFBF")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _conf_fill(v):
    return GREEN if v == "High" else (AMBER if v == "Medium" else "F2F2F2")


def _title(ws, text, span_col):
    ws["A1"] = text
    ws["A1"].font = Font(name=ARIAL, size=15, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(f"A1:{get_column_letter(span_col)}1")
    ws.row_dimensions[1].height = 30


def _headers(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEADER)
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        c.border = _border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22


def build_workbook(data, filename):
    themes = data["themes"]
    wb = Workbook()

    # Flatten quotes into evidence rows; remember each one's Raw Data row.
    ws1 = wb.active
    ws1.title = "Raw Data"
    _title(ws1, "Raw Data — Evidence Log (click a Theme to see how it was used)", 5)
    _headers(ws1, 3, ["Quote ID", "Participant", "Topic", "Verbatim quote", "Theme"],
             [12, 13, 16, 80, 42])
    ws2 = wb.create_sheet("Synthesis")
    ws3 = wb.create_sheet("Report")

    raw_row = {}        # (theme_idx, quote_idx) -> excel row
    r = 4
    for ti, t in enumerate(themes):
        for qi, q in enumerate(t["quotes"]):
            rid = f"{q['participant']}-{ti+1}{qi+1}"
            ws1.cell(row=r, column=1, value=rid).font = Font(name=ARIAL, size=10, bold=True)
            ws1.cell(row=r, column=2, value=q["participant"]).font = Font(name=ARIAL, size=10)
            ws1.cell(row=r, column=3, value=q.get("topic", "")).font = Font(name=ARIAL, size=10)
            ws1.cell(row=r, column=4, value=q["quote"]).font = Font(name=ARIAL, size=10)
            for col in range(1, 6):
                cc = ws1.cell(row=r, column=col)
                cc.border = _border
                cc.alignment = Alignment(vertical="top", wrap_text=True)
            ws1.row_dimensions[r].height = 46
            raw_row[(ti, qi)] = r
            r += 1
    ws1.freeze_panes = "A4"

    # Synthesis
    _title(ws2, "Synthesis — Themes & Evidence (click a quote to jump to its source)", 5)
    _headers(ws2, 3, ["Theme", "Confidence", "Supporting quote  (click to view source)",
                      "P#", "Recommendation"], [34, 13, 66, 7, 46])
    theme_first = {}
    r = 4
    for ti, t in enumerate(themes):
        theme_first[ti] = r
        for qi, q in enumerate(t["quotes"]):
            tn = ws2.cell(row=r, column=1, value=t["name"] if qi == 0 else "")
            tn.font = Font(name=ARIAL, size=10, bold=True)
            cf = ws2.cell(row=r, column=2, value=t["confidence"] if qi == 0 else "")
            cf.font = Font(name=ARIAL, size=10, bold=True)
            if qi == 0:
                cf.fill = PatternFill("solid", fgColor=_conf_fill(t["confidence"]))
            qc = ws2.cell(row=r, column=3, value=f'"{q["quote"]}"')
            qc.font = Font(name=ARIAL, size=10, color=LINK, underline="single")
            qc.hyperlink = Hyperlink(ref=qc.coordinate,
                                     location=f"'Raw Data'!A{raw_row[(ti, qi)]}",
                                     display="view source")
            ws2.cell(row=r, column=4, value=q["participant"]).font = Font(name=ARIAL, size=10)
            rc = ws2.cell(row=r, column=5, value=t["recommendation"] if qi == 0 else "")
            rc.font = Font(name=ARIAL, size=10)
            for col in range(1, 6):
                cc = ws2.cell(row=r, column=col)
                cc.border = _border
                cc.alignment = Alignment(vertical="top", wrap_text=True)
            ws2.row_dimensions[r].height = 44
            r += 1
    ws2.freeze_panes = "A4"

    # Raw Data Theme cells link up to Synthesis
    for ti, t in enumerate(themes):
        for qi, q in enumerate(t["quotes"]):
            c = ws1.cell(row=raw_row[(ti, qi)], column=5, value=t["name"])
            c.font = Font(name=ARIAL, size=10, color=LINK, underline="single")
            c.hyperlink = Hyperlink(ref=c.coordinate,
                                    location=f"'Synthesis'!A{theme_first[ti]}",
                                    display="view theme")

    # Report
    _title(ws3, f"Research Report — {data['product']}", 6)
    ws3["A2"] = f"Draft for review · {len(data['participants'])} participants · requires researcher verification"
    ws3["A2"].font = Font(name=ARIAL, size=10, italic=True, color="595959")
    ws3.merge_cells("A2:F2")
    _headers(ws3, 4, ["#", "Theme", "Confidence", "What we found", "Recommendation", "Evidence"],
             [5, 30, 13, 52, 46, 18])
    r = 5
    for ti, t in enumerate(themes):
        ws3.cell(row=r, column=1, value=ti + 1).font = Font(name=ARIAL, size=10, bold=True)
        ws3.cell(row=r, column=2, value=t["name"]).font = Font(name=ARIAL, size=10, bold=True)
        cf = ws3.cell(row=r, column=3, value=t["confidence"])
        cf.font = Font(name=ARIAL, size=10, bold=True)
        cf.fill = PatternFill("solid", fgColor=_conf_fill(t["confidence"]))
        ws3.cell(row=r, column=4, value=t["summary"]).font = Font(name=ARIAL, size=10)
        ws3.cell(row=r, column=5, value=t["recommendation"]).font = Font(name=ARIAL, size=10)
        ev = ws3.cell(row=r, column=6, value="see evidence →")
        ev.font = Font(name=ARIAL, size=10, color=LINK, underline="single")
        ev.hyperlink = Hyperlink(ref=ev.coordinate,
                                 location=f"'Synthesis'!A{theme_first[ti]}",
                                 display="see evidence")
        for col in range(1, 7):
            cc = ws3.cell(row=r, column=col)
            cc.border = _border
            cc.alignment = Alignment(vertical="top", wrap_text=True)
        ws3.row_dimensions[r].height = 58
        r += 1
    r += 1
    vh = ws3.cell(row=r, column=1, value="Needs human verification (before sharing these findings)")
    vh.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
    vh.fill = PatternFill("solid", fgColor="C00000")
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws3.row_dimensions[r].height = 22
    r += 1
    for v in data["verification_notes"]:
        vc = ws3.cell(row=r, column=1, value="•  " + v)
        vc.font = Font(name=ARIAL, size=10)
        vc.alignment = Alignment(vertical="top", wrap_text=True)
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws3.row_dimensions[r].height = 30
        r += 1
    ws3.freeze_panes = "A5"

    wb.save(filename)


# ─────────────────────────────────────────────────────────────────────────────
# 4. THE AGENT LOOP — think → act → observe → repeat.
# ─────────────────────────────────────────────────────────────────────────────

def run_agent(goal: str, client: anthropic.Anthropic, model: str = "claude-sonnet-4-6") -> str:
    messages = [{"role": "user", "content": goal}]
    while True:
        # ===== THIS IS THE API CALL — the one moment your code talks to Claude. =====
        # Everything gathered so far (instructions + transcripts read) is sent over the
        # internet to Anthropic; Claude thinks; the reply lands in `response`.
        # The loop runs this again every turn, so one agent run = several API calls.
        response = client.messages.create(
            model=model, max_tokens=4096, system=SYSTEM_PROMPT,
            tools=TOOLS, messages=messages,
        )
        messages.append({"role": "assistant", "content": response.content})
        if response.stop_reason != "tool_use":
            return "".join(b.text for b in response.content if b.type == "text")

        tool_results = []
        for block in response.content:
            if block.type == "tool_use":
                result = TOOL_FUNCTIONS[block.name](**block.input)
                name = block.name
                print(f"  [tool] {name} -> {str(result)[:70]}")
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": str(result),
                })
        messages.append({"role": "user", "content": tool_results})


if __name__ == "__main__":
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise SystemExit("Set ANTHROPIC_API_KEY first (see the top of this file).")
    # Opens the connection to Anthropic. It auto-reads your ANTHROPIC_API_KEY
    # (your password) from the environment and attaches it to every call below.
    client = anthropic.Anthropic()
    goal = "Synthesize all the interview transcripts and submit your findings."
    print(f"\nGOAL: {goal}\n")
    answer = run_agent(goal, client)
    print(f"\nAGENT:\n{answer}\n")
    print(f"Open {OUTPUT_FILE} to see the linked report.")
